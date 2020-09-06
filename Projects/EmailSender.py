import email, smtplib, ssl

from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from openpyxl import load_workbook


def sendMail(to):
    subject = "Invitation for Campus Recruitment | UIET Hoshiarpur"
    body = """<html>
        <body>
            <p>
    Sending thoughts of health and peace from UIET, Hoshiarpur!<br>
    COVID-19 had changed the standard hiring process, here’s my take on why you should be reading this. 
    <br>
    <br>
    Dear HR<br>
    During these uncertain times we would like to extend a solution which is beneficial and safe for both our organisations. This is not a pitch, but a prospect for your Company's success.
    <br>
    <br>
    <strong>Why hire from UIET, Hoshiarpur?</strong><br>
    University Institute of Engineering & Technology, Hoshiarpur is a part of an engineering institute of prestigious, Five stars accredited Panjab University, Chandigarh <a href="http://puchd.ac.in">(http://puchd.ac.in)</a>.<br>
    Panjab University is ranked among the top Universities in the world.<br>
    We have a batch of talented young minds, enthusiasts who don’t just want to limit their knowledge to textbooks but learn from real time projects. We have students willing to work for esteemed organisations like yours and they surely would not disappoint.<br>
    Our alumnus are placed in many reputed firms like Microsoft, Ericsson, Adobe, Airtel, Google, Swiggy, we even have alumnus who are entrepreneurs and also those who serve the nation with pride.
    <br>
    <br>
    <STRONG>What specializations does UIET, Hoshiarpur offer?</STRONG><br>
    UIET, Hoshiarpur offers B.E.(Bachelor of Engineering) degree in<br>
    <ul>
        <li>Electronics & Communication </li>
        <li>Computer Science </li>
        <li>Mechanical</li>
        <li>Information Technology</li>
    </ul>
    <br>
    We are industrious, willing to work in software development, data science, web development, networking, designing and much more!<br>
    <br>
    <STRONG>Ways of hiring</STRONG><br>
    We would like to invite your company for the placements and would be honored to be associated with your eminent organization.<br>
    We would suggest hiring to take place in the following way:
    <ul>
        <li>Online hiring methods</li>
        <li>Offline hiring methods</li>
    </ul><br>
    <i>For Final year students:</i>
    <ol>
        <li>Internship opportunity (Jan-May 2021)</li>
        <li>Full time opportunities (June 2021 onwards)</li>
        <li>Internship + Full time opportunities </li>
    </ol><br>
    <i>For Prefinal year students:</i>
    <ol>
        <li>Summer Internship opportunity (June-July 2021)</li>
    </ol><br>
    We firmly believe that your organization and our institute can endeavor to gain immensely from this symbiotic relationship.<br><br>
    Warm Regards<br>
    Mr. Rajeev Kumar Dang<br>
    Placement Coordinator<br>
    University Institute Of Engineering & Technology<br>
    Panjab University SSG Regional Centre, Hoshiarpur<br>
    M.No.: (+91-********)<br>
    Campus Website- <a href="http://ssgpurch.puchd.ac.in">http://ssgpurch.puchd.ac.in</a>
    <br><br>
    Nakul Arya<br>
    Student Coordinator<br>
    Training and Placement Cell<br>
    (+91-**************)<br>
            </p>
        </body>
    </html>"""
    sender_email = "nakularya99@gmail.com"
    receiver_email = to
    password = "********"

    # Create a multipart message and set headers
    message = MIMEMultipart()
    message["From"] = sender_email
    message["To"] = receiver_email
    message["Subject"] = subject
    # message["Bcc"] = receiver_email  # Recommended for mass emails

    # Add body to email
    message.attach(MIMEText(body, "html"))

    filename = "UIETH PlacementBrochure 20-21.pdf"  # In same directory as script

    # Open PDF file in binary mode
    with open(filename, "rb") as attachment:
        # Add file as application/octet-stream
        # Email client can usually download this automatically as attachment
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())

    # Encode file in ASCII characters to send by email
    encoders.encode_base64(part)

    # Add header as key/value pair to attachment part
    part.add_header(
        "Content-Disposition",
        f"attachment; filename= {filename}",
    )

    # Add attachment to message and convert message to string
    message.attach(part)
    text = message.as_string()

    # Log in to server using secure context and send email
    context = ssl.create_default_context()
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
        try:
            server.login(sender_email, password)
            server.sendmail(sender_email, receiver_email, text)
        except Exception as e:
            print(e)


count = 1449
wb = load_workbook(filename='HRS.xlsx')
ws = wb['Sheet1']
for rowCount in range(count, 1700):
    sendMail(ws.cell(row=rowCount, column=3).value)
    print(count)
    count += 1
